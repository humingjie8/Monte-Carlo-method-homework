import numpy as np
import matplotlib.pyplot as plt
import math
import seaborn as sns
import openpyxl
from tqdm import tqdm

# 使用切片来获取前N项并计算它们的和
def sum_of_first_n_elements(arr, n):
    return sum(arr[:n])


#设置总样本数据量
number_data_z = 0
number_data_t = 1000
#set for data cycle number
cycle = 100

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()

# 创建一个工作表（默认名称是Sheet）
sheet = workbook.active
sheet.cell(row=1, column=1, value='x')
sheet.cell(row=1, column=2, value=0.01)
sheet.cell(row=1, column=3, value=0.05)
sheet.cell(row=1, column=4, value=0.1)

t = [0] * number_data_t
for i in tqdm(range(cycle), desc="Processing"):
    number_data_z += 100
    for x in range(number_data_t):
        seed = x
        # 设置随机数种子以确保可重复性
        np.random.seed(seed)

        # 生成10000个参数为1的指数分布数据点
        lambda_parameter = 1.0  # 参数λ为1
        data = np.random.exponential(scale=1.0 / lambda_parameter, size=number_data_z)
        mean_value = np.mean(data)
        min_value = np.min(data)



        # for i in range(number_data):
        mid_1 = sum_of_first_n_elements((data-mean_value)**2, number_data_z)
        t[x] = (math.sqrt(mid_1/(number_data_z)))/(mean_value-min_value)
        # print(np.min(t))

    # 计算分位数
    N1 = np.percentile(t, 0.01)
    N2 = np.percentile(t, 0.05)
    N3 = np.percentile(t, 0.1)
    # print(N1, N2, N3)

    # 在EXCEL中写入数据
    sheet.cell(row=i+2, column=1, value=number_data_z)  # 在第一列写入数据
    sheet.cell(row=i+2, column=2, value=str(N1))  # 在第2列写入数据    
    sheet.cell(row=i+2, column=3, value=str(N2))  # 在第3列写入数据
    sheet.cell(row=i+2, column=4, value=str(N3))  # 在第4列写入数据

    # 绘制KDE图
    label = 'number_z='+str(number_data_z)
    sns.kdeplot(t, fill=False, label=label, color='r')
    plt.title('Kernel Density Estimation (KDE) of Data')

    # 设置横坐标和纵坐标的范围
    x_max = 1.5
    # if number_data > 100 :
    #     x_max = 0.5
    plt.xlim(0.5, x_max)  # 设置横坐标的范围为1到5
    # plt.ylim(0, 12)  # 设置纵坐标的范围为0到12

    plt.xlabel('Value')
    plt.ylabel('Density')
    plt.legend()

    name_pic = 'ex_pic/'+'ex'+str(number_data_z)+'.png'
    plt.savefig(name_pic)

    # plt.show()
    plt.clf()

# 保存Excel文件
name_excel = 'ex'+'.xlsx'
workbook.save(name_excel)

# name_pic = 'ex_pic/'+'ex'+'.png'
# plt.savefig(name_pic)

# 关闭工作簿
workbook.close()
